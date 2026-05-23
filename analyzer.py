import pandas as pd
import numpy as np
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = "data"
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# regex patterns for German/EU formats
PATTERN_DE_POSTAL = re.compile(r"^\d{5}$")
PATTERN_VAT_DE = re.compile(r"^DE\d{9}$")
PATTERN_IBAN = re.compile(r"^[A-Z]{2}\d{2}[A-Z0-9]{11,30}$")
PATTERN_EMAIL = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
PATTERN_PHONE = re.compile(r"^\+?\d[\d\s\-/]{6,20}$")

# SAP requirements per entity (mandatory fields cannot be blank)
SAP_RULES = {
    "customers": {
        "mandatory": ["COMPANY_NAME", "STREET", "POSTAL_CODE", "CITY", "COUNTRY"],
        "format_checks": {
            "POSTAL_CODE": (PATTERN_DE_POSTAL, "Must be 5 digits (German postal code)"),
            "VAT_ID": (PATTERN_VAT_DE, "Must be DE + 9 digits, e.g. DE123456789"),
            "IBAN": (PATTERN_IBAN, "Invalid IBAN format"),
            "EMAIL": (PATTERN_EMAIL, "Invalid email format"),
        },
        "duplicate_key": "COMPANY_NAME",
    },
    "suppliers": {
        "mandatory": ["COMPANY_NAME", "STREET", "POSTAL_CODE", "CITY", "COUNTRY"],
        "format_checks": {
            "POSTAL_CODE": (PATTERN_DE_POSTAL, "Must be 5 digits"),
            "VAT_ID": (PATTERN_VAT_DE, "Must be DE + 9 digits"),
            "IBAN": (PATTERN_IBAN, "Invalid IBAN format"),
            "EMAIL": (PATTERN_EMAIL, "Invalid email format"),
        },
        "duplicate_key": "COMPANY_NAME",
    },
    "materials": {
        "mandatory": ["MATERIAL_NUMBER", "DESCRIPTION", "UNIT", "MATERIAL_GROUP"],
        "format_checks": {},
        "duplicate_key": "DESCRIPTION",
    },
    "open_items": {
        "mandatory": ["DOCUMENT_NUMBER", "TYPE", "PARTNER_ID", "POSTING_DATE",
                      "DUE_DATE", "GROSS_AMOUNT", "CURRENCY"],
        "format_checks": {},
        "duplicate_key": "DOCUMENT_NUMBER",
    },
}


# issue severities:
# CRITICAL: missing mandatory field (fails SAP import)
# WARNING: format violation (SAP might misread or reject)
# INFO: duplicate found (needs manual check)

def analyze_entity(df, entity_name):
    rules = SAP_RULES[entity_name]
    issues = []

    for idx, row in df.iterrows():
        row_issues = []

        # 1. check mandatory fields
        for field in rules["mandatory"]:
            if field not in df.columns:
                continue
            val = row.get(field, "")
            if pd.isna(val) or str(val).strip() == "":
                row_issues.append({
                    "field": field,
                    "severity": "CRITICAL",
                    "message": f"Mandatory field '{field}' is empty — SAP import will fail",
                })

        # 2. check formats (only if filled)
        for field, (pattern, hint) in rules["format_checks"].items():
            if field not in df.columns:
                continue
            val = str(row.get(field, "")).strip()
            if val == "" or pd.isna(row.get(field)):
                continue  # already caught above if mandatory

            if not pattern.match(val):
                row_issues.append({
                    "field": field,
                    "severity": "WARNING",
                    "message": f"'{val}' — {hint}",
                })

        for issue in row_issues:
            issues.append({
                "record_index": idx,
                "legacy_id": row.get(list(df.columns)[0], idx),
                "company_or_id": row.get("COMPANY_NAME", row.get("DESCRIPTION",
                                                                 row.get("DOCUMENT_NUMBER", str(idx)))),
                **issue,
            })

    # 3. check duplicates
    dup_key = rules.get("duplicate_key")
    if dup_key and dup_key in df.columns:
        dup_mask = df[dup_key].str.strip().str.lower().duplicated(keep=False)
        dup_vals = df.loc[dup_mask, dup_key].unique()

        for val in dup_vals:
            dup_rows = df[df[dup_key].str.strip().str.lower() == val.strip().lower()]
            ids = ", ".join(str(x) for x in dup_rows.iloc[:, 0].tolist())
            issues.append({
                "record_index": -1,
                "legacy_id": ids,
                "company_or_id": val,
                "field": dup_key,
                "severity": "INFO",
                "message": f"Duplicate value '{val}' found in records: {ids}",
            })

    issues_df = pd.DataFrame(issues) if issues else pd.DataFrame(
        columns=["record_index", "legacy_id", "company_or_id", "field", "severity", "message"]
    )

    # calc readiness score (100 minus penalty)
    total = len(df)
    penalty = 0
    penalty += issues_df[issues_df["severity"] == "CRITICAL"].shape[0] * 5
    penalty += issues_df[issues_df["severity"] == "WARNING"].shape[0] * 2
    penalty += issues_df[issues_df["severity"] == "INFO"].shape[0] * 1

    max_penalty = total * 5
    score = max(0, round(100 - (penalty / max_penalty * 100) if max_penalty > 0 else 100, 1))

    return issues_df, score


# excel export colors
RED = "FFCCCC"
YELLOW = "FFF3CC"
BLUE = "CCE5FF"
GREEN = "CCFFCC"
HEADER = "1F3864"
WHITE = "FFFFFF"


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _header_font():
    return Font(bold=True, color=WHITE, size=10)


def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def build_excel_report(results):
    """
    results: dict of {entity_name: (issues_df, score, total_records)}
    """
    output_path = f"{OUTPUT_DIR}/Migration_Readiness_Report.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # build summary sheet
        summary_rows = []
        for entity, (issues_df, score, total) in results.items():
            critical = issues_df[issues_df["severity"] == "CRITICAL"].shape[0] if len(issues_df) else 0
            warnings = issues_df[issues_df["severity"] == "WARNING"].shape[0] if len(issues_df) else 0
            info = issues_df[issues_df["severity"] == "INFO"].shape[0] if len(issues_df) else 0
            status = "READY" if score >= 85 else ("REVIEW NEEDED" if score >= 65 else "NOT READY")

            summary_rows.append({
                "Entity": entity.replace("_", " ").title(),
                "Total Records": total,
                "Readiness Score": f"{score}%",
                "Critical Issues": critical,
                "Warnings": warnings,
                "Info / Duplicates": info,
                "Go-Live Status": status,
            })

        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        ws = writer.sheets["Summary"]
        ws.insert_rows(1, 3)
        ws["B2"] = "ERP Migration Readiness Report"
        ws["B2"].font = Font(bold=True, size=14, color=HEADER)
        ws["B3"] = "Hoffmann & Partner Maschinenbau GmbH  |  Target: SAP S/4HANA Cloud / SAP Business ByDesign"
        ws["B3"].font = Font(italic=True, size=10, color="555555")

        header_row = 4
        for col_idx, col_name in enumerate(summary_df.columns, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill = _fill(HEADER)
            cell.font = _header_font()
            cell.alignment = Alignment(horizontal="center")

        for row_idx in range(header_row + 1, header_row + 1 + len(summary_df)):
            score_val = ws.cell(row=row_idx, column=3).value
            score_num = float(str(score_val).replace("%", "")) if score_val else 0

            color = GREEN if score_num >= 85 else (YELLOW if score_num >= 65 else RED)
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = _fill(color)
                cell.border = _border()
                cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 35)

        # detail sheets per entity
        for entity, (issues_df, score, total) in results.items():
            sheet_name = entity.replace("_", " ").title()[:31]
            if len(issues_df) == 0:
                pd.DataFrame([{"Status": "No issues found. Entity is ready for migration."}]) \
                    .to_excel(writer, sheet_name=sheet_name, index=False)
                continue

            issues_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws_detail = writer.sheets[sheet_name]

            for row_idx, (_, issue_row) in enumerate(issues_df.iterrows(), start=2):
                sev = issue_row.get("severity", "")
                color = RED if sev == "CRITICAL" else (YELLOW if sev == "WARNING" else BLUE)
                for col_idx in range(1, len(issues_df.columns) + 1):
                    cell = ws_detail.cell(row=row_idx, column=col_idx)
                    cell.fill = _fill(color)
                    cell.border = _border()

            for col_idx in range(1, len(issues_df.columns) + 1):
                cell = ws_detail.cell(row=1, column=col_idx)
                cell.fill = _fill(HEADER)
                cell.font = _header_font()
                cell.alignment = Alignment(horizontal="center")

            for col in ws_detail.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws_detail.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    return output_path


def run_analysis():
    entities = {
        "customers": pd.read_csv(f"{DATA_DIR}/customers.csv", dtype=str).fillna(""),
        "suppliers": pd.read_csv(f"{DATA_DIR}/suppliers.csv", dtype=str).fillna(""),
        "materials": pd.read_csv(f"{DATA_DIR}/materials.csv", dtype=str).fillna(""),
        "open_items": pd.read_csv(f"{DATA_DIR}/open_items.csv", dtype=str).fillna(""),
    }

    results = {}
    print("\nRunning validation checks...\n")

    for name, df in entities.items():
        issues_df, score = analyze_entity(df, name)
        results[name] = (issues_df, score, len(df))

        critical = issues_df[issues_df["severity"] == "CRITICAL"].shape[0] if len(issues_df) else 0
        warnings = issues_df[issues_df["severity"] == "WARNING"].shape[0] if len(issues_df) else 0
        info = issues_df[issues_df["severity"] == "INFO"].shape[0] if len(issues_df) else 0
        status = "READY" if score >= 85 else ("REVIEW NEEDED" if score >= 65 else "NOT READY")

        print(f"  {name.upper():<12} | {len(df):>4} records | Score: {score:>5}% | "
              f"Critical: {critical:>3} | Warnings: {warnings:>3} | Info: {info:>3} | {status}")

    report_path = build_excel_report(results)

    total_issues = sum(len(r[0]) for r in results.values())
    total_records = sum(r[2] for r in results.values())
    avg_score = round(sum(r[1] for r in results.values()) / len(results), 1)

    print(f"\n{'=' * 60}")
    print(f"  Total records checked:   {total_records}")
    print(f"  Total issues found:      {total_issues}")
    print(f"  Average readiness score: {avg_score}%")
    print(f"  Report saved to:         {report_path}")
    print(f"{'=' * 60}")

    return results, avg_score